#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mitarbeiter-Jahres-Excel-Export fuer TeamFlow
Erstellt eine formatierte Excel-Datei mit Jahresuebersicht eines Mitarbeiters
"""

import sys
import json
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FEHLER: openpyxl nicht installiert!", file=sys.stderr)
    print("Installiere mit: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Farben ────────────────────────────────────────────────────────────────────
C_PRIMARY_BG  = "1F538D"
C_PRIMARY_FG  = "FFFFFF"
C_TITLE_FONT  = "1F538D"
C_URLAUB      = "D6F0D6"
C_KRANKHEIT   = "FAD4D4"
C_SCHULUNG    = "D4EEF7"
C_UEBERSTD    = "FFF3CD"
C_SUMME_BG    = "E8E8E8"
C_SECTION_BG  = "EEF3FA"
C_GREY_BG     = "F5F5F5"

TYP_FARBEN = {
    "urlaub":       C_URLAUB,
    "krankheit":    C_KRANKHEIT,
    "schulung":     C_SCHULUNG,
    "ueberstunden": C_UEBERSTD,
}

TYP_LABEL = {
    "urlaub":       "Urlaub",
    "krankheit":    "Krankheit",
    "schulung":     "Schulung",
    "ueberstunden": "Ueberstunden",
}


def make_border(thin=True):
    s = "thin" if thin else "medium"
    side = Side(style=s)
    return Border(left=side, right=side, top=side, bottom=side)


def fmt_datum(d):
    if not d:
        return ""
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d


def fmt_zahl(v, einheit=""):
    if v is None:
        return "–"
    try:
        f = float(v)
        zahl = int(f) if f == int(f) else round(f, 2)
        return f"{zahl}{einheit}"
    except Exception:
        return "–"


def style_header(cell, bg=C_PRIMARY_BG, fg=C_PRIMARY_FG, size=10, bold=True, center=True):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True
    )
    cell.border = make_border()


def style_data(cell, bg=None, center=False, bold=False, size=9):
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center"
    )
    cell.border = make_border()


def style_label(cell, text, bold=True, size=9, color="666666"):
    cell.value = text
    cell.font = Font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = make_border()


def style_value(cell, text, bold=False, size=10):
    cell.value = text
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = make_border()


# ── Tabellenblatt 1: Uebersicht ───────────────────────────────────────────────
def schreibe_uebersicht(wb, emp, jahr, stats):
    ws = wb.active
    ws.title = "Uebersicht"

    name = emp.get("name", "–")
    department = emp.get("department", "–")

    # Titelzeile
    ws.merge_cells("A1:F1")
    tc = ws["A1"]
    tc.value = f"Jahresuebersicht {jahr}  –  {name}"
    tc.font = Font(color=C_TITLE_FONT, bold=True, size=16)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Untertitel
    ws.merge_cells("A2:F2")
    sc = ws["A2"]
    sc.value = f"{department}  |  Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sc.font = Font(color="888888", italic=True, size=9)
    sc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8  # Abstandszeile

    # ── Urlaubsstatistik ──────────────────────────────────────────────────────
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    hc = ws.cell(row=row, column=1, value="URLAUBSSTATISTIK")
    hc.font = Font(color=C_TITLE_FONT, bold=True, size=11)
    hc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    stat_felder = [
        ("Jahresanspruch",              fmt_zahl(stats.get("urlaubsanspruch"), " Tage")),
        (f"Uebertrag aus {int(jahr)-1}", fmt_zahl(stats.get("uebertrag_vorjahr"), " Tage")),
        ("Verfuegbar gesamt",           fmt_zahl(stats.get("urlaub_verfuegbar"), " Tage")),
        ("Genommen",                    fmt_zahl(stats.get("urlaub_genommen"), " Tage")),
        ("Resturlaub",                  fmt_zahl(stats.get("urlaub_rest"), " Tage")),
    ]

    for i, (label, wert) in enumerate(stat_felder):
        bg = C_SUMME_BG if label == "Resturlaub" else (C_GREY_BG if i % 2 == 0 else None)
        bold_row = label == "Resturlaub"

        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=bold_row, color="666666", size=9)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = make_border()
        if bg:
            lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        ws.merge_cells(f"B{row}:F{row}")
        vc = ws.cell(row=row, column=2, value=wert)
        vc.font = Font(bold=bold_row, size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = make_border()
        if bg:
            vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        # Merge-Zellen Border-Fix
        for col in range(3, 7):
            c = ws.cell(row=row, column=col)
            c.border = make_border()
            if bg:
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        ws.row_dimensions[row].height = 18
        row += 1

    # ── Weitere Abwesenheiten ─────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    hc2 = ws.cell(row=row, column=1, value="WEITERE ABWESENHEITEN")
    hc2.font = Font(color=C_TITLE_FONT, bold=True, size=11)
    hc2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    weitere = [
        ("Krankheitstage",      fmt_zahl(stats.get("krankheitstage"), " Tage")),
        ("Schulungstage",       fmt_zahl(stats.get("schulungstage"), " Tage")),
        ("Ueberstunden-Saldo",  fmt_zahl(stats.get("ueberstunden"), "h")),
    ]

    for i, (label, wert) in enumerate(weitere):
        bg = C_GREY_BG if i % 2 == 0 else None

        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(color="666666", size=9)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = make_border()
        if bg:
            lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        ws.merge_cells(f"B{row}:F{row}")
        vc = ws.cell(row=row, column=2, value=wert)
        vc.font = Font(size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = make_border()
        if bg:
            vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col in range(3, 7):
            c = ws.cell(row=row, column=col)
            c.border = make_border()
            if bg:
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        ws.row_dimensions[row].height = 18
        row += 1

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 10

    ws.freeze_panes = "A4"


# ── Tabellenblatt 2: Eintraege ────────────────────────────────────────────────
def schreibe_eintraege(wb, emp, jahr, eintraege):
    ws = wb.create_sheet("Eintraege")

    name = emp.get("name", "–")
    department = emp.get("department", "–")

    # Titel
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value = f"Eintraege {jahr}  –  {name}"
    tc.font = Font(color=C_TITLE_FONT, bold=True, size=14)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sc = ws["A2"]
    sc.value = f"{department}  |  Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sc.font = Font(color="888888", italic=True, size=9)
    sc.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8

    # Header
    headers = ["Typ", "Von", "Bis", "Wert", "Titel / Notiz", "Notiz", ""]
    # Vereinfachter Header: Typ | Von | Bis | Wert | Titel | Notiz
    headers = ["Typ", "Von", "Bis", "Wert", "Titel", "Notiz"]
    col_widths = [16, 14, 14, 12, 25, 35]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[4].height = 22

    if not eintraege:
        ws.merge_cells("A5:F5")
        nc = ws.cell(row=5, column=1, value="Keine Eintraege vorhanden.")
        nc.font = Font(color="888888", italic=True, size=9)
        nc.alignment = Alignment(horizontal="center")
    else:
        # Typ-Trennzeilen
        aktueller_typ = None
        data_row = 5

        for eintrag in eintraege:
            typ = eintrag.get("typ", "")
            farbe = TYP_FARBEN.get(typ, "FFFFFF")
            label = TYP_LABEL.get(typ, typ)

            # Neue Typ-Gruppe → Trennzeile
            if typ != aktueller_typ:
                aktueller_typ = typ
                ws.merge_cells(f"A{data_row}:F{data_row}")
                gc = ws.cell(row=data_row, column=1, value=label.upper())
                gc.fill = PatternFill(start_color=C_PRIMARY_BG, end_color=C_PRIMARY_BG, fill_type="solid")
                gc.font = Font(color=C_PRIMARY_FG, bold=True, size=10)
                gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                gc.border = make_border()
                for col in range(2, 7):
                    c = ws.cell(row=data_row, column=col)
                    c.fill = PatternFill(start_color=C_PRIMARY_BG, end_color=C_PRIMARY_BG, fill_type="solid")
                    c.border = make_border()
                ws.row_dimensions[data_row].height = 20
                data_row += 1

            von_str = fmt_datum(eintrag.get("von_datum"))
            bis_str = fmt_datum(eintrag.get("bis_datum") or eintrag.get("von_datum"))
            wert = eintrag.get("wert", 0)

            if typ == "ueberstunden":
                try:
                    f = float(wert)
                    vorzeichen = "+" if f >= 0 else ""
                    wert_str = f"{vorzeichen}{fmt_zahl(wert)}h"
                except Exception:
                    wert_str = f"{wert}h"
            else:
                wert_str = fmt_zahl(wert, " T")

            titel = eintrag.get("titel") or ""
            notiz = eintrag.get("notiz") or ""

            zeile = [label, von_str, bis_str, wert_str, titel, notiz]

            for col, val in enumerate(zeile, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                center = col in (2, 3, 4)
                style_data(cell, bg=farbe, center=center)

            ws.row_dimensions[data_row].height = 16
            data_row += 1

    # Spaltenbreiten
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


# ── Tabellenblatt 3: Legende ──────────────────────────────────────────────────
def schreibe_legende(wb):
    ws = wb.create_sheet("Legende")

    ws["A1"].value = "Farbcode"
    style_header(ws["A1"])
    ws["B1"].value = "Bedeutung"
    style_header(ws["B1"])
    ws.row_dimensions[1].height = 20

    legende = [
        (C_URLAUB,     "Urlaub (Tage)"),
        (C_KRANKHEIT,  "Krankheit (Tage)"),
        (C_SCHULUNG,   "Schulung (Tage)"),
        (C_UEBERSTD,   "Ueberstunden (Stunden)"),
        (C_SUMME_BG,   "Summenzeile / Resturlaub"),
    ]

    for i, (farbe, text) in enumerate(legende, 2):
        ca = ws.cell(row=i, column=1, value="")
        ca.fill = PatternFill(start_color=farbe, end_color=farbe, fill_type="solid")
        ca.border = make_border()
        cb = ws.cell(row=i, column=2, value=text)
        style_data(cb)
        ws.row_dimensions[i].height = 18

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30


# ── Haupt ─────────────────────────────────────────────────────────────────────
def create_employee_year_excel(data, output_path):
    emp      = data.get("employee", {})
    jahr     = data.get("jahr", str(datetime.now().year))
    stats    = data.get("stats", {})
    eintraege = data.get("eintraege", [])

    wb = Workbook()

    schreibe_uebersicht(wb, emp, jahr, stats)
    schreibe_eintraege(wb, emp, jahr, eintraege)
    schreibe_legende(wb)

    wb.save(output_path)
    sys.stdout.buffer.write(f"Excel erfolgreich erstellt: {output_path}\n".encode("utf-8"))


def main():
    if len(sys.argv) != 3:
        sys.stderr.write(
            "FEHLER: Usage: python export_employee_year_excel.py <input.json> <output.xlsx>\n"
        )
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]

    try:
        with io.open(input_file, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
        sys.stdout.buffer.write(b"JSON gelesen\n")
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Lesen der JSON: {e}\n".encode("utf-8"))
        sys.exit(1)

    try:
        create_employee_year_excel(data, output_file)
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Erstellen der Excel: {e}\n".encode("utf-8"))
        sys.exit(1)


if __name__ == "__main__":
    main()
