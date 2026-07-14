#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kalender-Export fuer TeamFlow (Excel)
Tagesweise Uebersicht aller Abwesenheiten + Liste der Tage,
an denen alle Mitarbeiter anwesend waren (ohne Wochenende/Feiertag).
"""

import sys
import json
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FEHLER: openpyxl nicht installiert!", file=sys.stderr)
    print("Installiere mit: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Farben ──────────────────────────────────────────────────────────────────
C_HEADER_BG     = "1F538D"
C_HEADER_FONT   = "FFFFFF"
C_TITLE_FONT    = "1F538D"
C_ALLE_DA_BG    = "D6F0D6"
C_WOCHENENDE_BG = "E8E8E8"
C_FEIERTAG_BG   = "E4D6F5"

TYP_LABEL = {
    "urlaub":       "Urlaub",
    "krankheit":    "Krankheit",
    "schulung":     "Schulung",
    "ueberstunden": "Ueberstunden-Abbau",
}


def make_border():
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)


def fmt_datum(d):
    if not d:
        return ""
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d


def fmt_zahl(v):
    if v is None:
        return 0
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except Exception:
        return 0


def style_header_cell(cell, bg=C_HEADER_BG, fg=C_HEADER_FONT):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = make_border()


def create_excel(payload, output_path):
    export_data   = payload.get("exportData", payload)
    tage          = export_data.get("tage", [])
    alle_anwesend = export_data.get("alleAnwesendTage", [])
    von_datum     = export_data.get("vonDatum", "")
    bis_datum     = export_data.get("bisDatum", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Kalenderuebersicht"

    # Titel
    ws.merge_cells("A1:D1")
    titel = ws["A1"]
    titel.value = f"Kalenderuebersicht  |  {fmt_datum(von_datum)} - {fmt_datum(bis_datum)}"
    titel.font = Font(color=C_TITLE_FONT, bold=True, size=14)
    titel.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(color="888888", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 16

    row = 4

    # ── Abschnitt: Alle Mitarbeiter anwesend ──
    ws.merge_cells(f"A{row}:D{row}")
    ueberschrift = ws.cell(row=row, column=1, value=f"Alle Mitarbeiter anwesend ({len(alle_anwesend)} Tage)")
    ueberschrift.font = Font(bold=True, size=11, color=C_TITLE_FONT)
    ueberschrift.fill = PatternFill(start_color=C_ALLE_DA_BG, end_color=C_ALLE_DA_BG, fill_type="solid")
    ueberschrift.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    if alle_anwesend:
        text = ", ".join(f"{t.get('wochentagName','')[:2]}. {fmt_datum(t.get('datum'))}" for t in alle_anwesend)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        ws.row_dimensions[row].height = max(18, 14 * (len(text) // 110 + 1))
    else:
        cell = ws.cell(row=row, column=1, value="Keine Tage mit vollstaendiger Anwesenheit im Zeitraum")
        cell.font = Font(size=9, italic=True, color="888888")
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row += 2  # + Leerzeile

    # ── Header Kalender-Tabelle ──
    headers = ["Datum", "Wochentag", "Status", "Eintraege"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        style_header_cell(cell)
    ws.row_dimensions[row].height = 20
    header_row = row
    row += 1

    for tag in tage:
        datum_str      = fmt_datum(tag.get("datum"))
        wochentag      = tag.get("wochentagName", "")
        ist_wochenende = tag.get("istWochenende", False)
        ist_feiertag   = tag.get("istFeiertag", False)
        feiertag_name  = tag.get("feiertagName") or ""
        alle_da        = tag.get("alleAnwesend", False)
        eintraege      = tag.get("eintraege", [])

        if alle_da:
            status, status_bg = "Alle anwesend", C_ALLE_DA_BG
        elif ist_wochenende:
            status, status_bg = "Wochenende", C_WOCHENENDE_BG
        elif ist_feiertag:
            status, status_bg = f"Feiertag: {feiertag_name}", C_FEIERTAG_BG
        else:
            status, status_bg = "", None

        if eintraege:
            zeilen = []
            for e in eintraege:
                wert = fmt_zahl(e.get("wert", 0))
                einheit = e.get("einheit", "T")
                typ_label = TYP_LABEL.get(e.get("typ", ""), e.get("typ", ""))
                notiz = e.get("notiz") or ""
                zusatz = f" - {notiz}" if notiz else ""
                zeilen.append(f"{e.get('mitarbeiter','')}: {typ_label} ({wert} {einheit}){zusatz}")
            eintraege_text = "\n".join(zeilen)
        else:
            eintraege_text = ""

        werte = [datum_str, wochentag, status, eintraege_text]
        for col, wert in enumerate(werte, 1):
            cell = ws.cell(row=row, column=col, value=wert)
            cell.font = Font(size=9)
            cell.alignment = Alignment(
                horizontal='center' if col in (1, 2, 3) else 'left',
                vertical='center',
                wrap_text=(col == 4)
            )
            cell.border = make_border()
            if col == 3 and status_bg:
                cell.fill = PatternFill(start_color=status_bg, end_color=status_bg, fill_type="solid")

        anzahl_zeilen = max(1, len(eintraege))
        ws.row_dimensions[row].height = 16 * anzahl_zeilen if eintraege else 16
        row += 1

    # Spaltenbreiten (unabhaengig von gemergten Zellen in Zeile 1/2!)
    breiten = [14, 14, 20, 60]
    for i, b in enumerate(breiten, 1):
        ws.column_dimensions[get_column_letter(i)].width = b

    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(output_path)
    sys.stdout.buffer.write(f"Excel erfolgreich erstellt: {output_path}\n".encode("utf-8"))


def main():
    if len(sys.argv) != 3:
        sys.stderr.write("FEHLER: Usage: python export_kalender_excel.py <input.json> <output.xlsx>\n")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]

    try:
        with io.open(input_file, "r", encoding="utf-8-sig") as f:
            payload = json.load(f)
        sys.stdout.buffer.write(b"JSON gelesen\n")
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Lesen der JSON: {e}\n".encode("utf-8"))
        sys.exit(1)

    try:
        create_excel(payload, output_file)
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Erstellen der Excel: {e}\n".encode("utf-8"))
        sys.exit(1)


if __name__ == "__main__":
    main()
