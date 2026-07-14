#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ueberstunden-Export fuer TeamFlow (Excel)
Nur geleistete Ueberstunden (Stunden > 0): wer, wann, wie viel, Notiz + Summen.
"""

import sys
import json
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FEHLER: openpyxl nicht installiert!", file=sys.stderr)
    print("Installiere mit: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Farben ──────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F538D"
C_HEADER_FONT = "FFFFFF"
C_ABT_BG     = "2D5FA8"
C_ABT_FONT   = "FFFFFF"
C_UEBERSTD_G = "CFE2FF"
C_SUMME_BG   = "E8E8E8"
C_TITLE_FONT = "1F538D"


def make_border():
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)


def fmt_datum(d):
    if not d:
        return ""
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d


def fmt_zahl(v):
    if v is None:
        return 0
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except Exception:
        return 0


def style_header_cell(cell, bg=C_HEADER_BG, fg=C_HEADER_FONT):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = make_border()


def style_data_cell(cell, bg=None, center=False, bold=False):
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(bold=bold, size=9)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center'
    )
    cell.border = make_border()


# ── Tabellenblatt 1: Zusammenfassung je Mitarbeiter ─────────────────────────
def schreibe_zusammenfassung(wb, mitarbeiter_liste, von_datum, bis_datum, gesamt_alle):
    ws = wb.active
    ws.title = "Zusammenfassung"

    ws.merge_cells("A1:D1")
    titel_cell = ws["A1"]
    titel_cell.value = f"Ueberstunden-Uebersicht  |  {fmt_datum(von_datum)} - {fmt_datum(bis_datum)}"
    titel_cell.font = Font(color=C_TITLE_FONT, bold=True, size=14)
    titel_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(color="888888", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8

    headers = ["Mitarbeiter", "Abteilung", "Geleistete Stunden", "Aktueller Saldo", "Anzahl Eintraege"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        style_header_cell(cell)
    ws.row_dimensions[4].height = 22

    row = 5
    aktuelle_abteilung = None

    for eintrag in mitarbeiter_liste:
        ma = eintrag.get("mitarbeiter", {})
        abt = ma.get("abteilung", "")
        name = ma.get("name", "")

        if abt != aktuelle_abteilung:
            aktuelle_abteilung = abt
            ws.merge_cells(f"A{row}:E{row}")
            abt_cell = ws.cell(row=row, column=1, value=abt)
            abt_cell.fill = PatternFill(start_color=C_ABT_BG, end_color=C_ABT_BG, fill_type="solid")
            abt_cell.font = Font(color=C_ABT_FONT, bold=True, size=10)
            abt_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            abt_cell.border = make_border()
            ws.row_dimensions[row].height = 20
            row += 1

        werte = [
            name,
            abt,
            fmt_zahl(eintrag.get("gesamtStunden", 0)),
            fmt_zahl(eintrag.get("aktuellerSaldo", 0)),
            len(eintrag.get("eintraege", [])),
        ]
        for col, wert in enumerate(werte, 1):
            cell = ws.cell(row=row, column=col, value=wert)
            style_data_cell(cell, center=col > 2)
        ws.row_dimensions[row].height = 18
        row += 1

    # Summenzeile
    ws.merge_cells(f"A{row}:B{row}")
    summe_cell = ws.cell(row=row, column=1, value="GESAMT")
    style_data_cell(summe_cell, bg=C_SUMME_BG, bold=True, center=True)
    ws.cell(row=row, column=2).fill = PatternFill(start_color=C_SUMME_BG, end_color=C_SUMME_BG, fill_type="solid")
    ws.cell(row=row, column=2).border = make_border()

    cell_c = ws.cell(row=row, column=3, value=fmt_zahl(gesamt_alle))
    style_data_cell(cell_c, bg=C_SUMME_BG, bold=True, center=True)
    cell_d = ws.cell(row=row, column=4, value=fmt_zahl(sum(fmt_zahl(e.get("aktuellerSaldo", 0)) for e in mitarbeiter_liste)))
    style_data_cell(cell_d, bg=C_SUMME_BG, bold=True, center=True)
    cell_e = ws.cell(row=row, column=5, value=sum(len(e.get("eintraege", [])) for e in mitarbeiter_liste))
    style_data_cell(cell_e, bg=C_SUMME_BG, bold=True, center=True)
    ws.row_dimensions[row].height = 20

    breiten = [28, 20, 18, 16, 18]
    for i, b in enumerate(breiten, 1):
        ws.column_dimensions[get_column_letter(i)].width = b

    ws.freeze_panes = "A5"


# ── Tabellenblatt 2: Detailtabelle ──────────────────────────────────────────
def schreibe_detail(wb, mitarbeiter_liste, von_datum, bis_datum):
    ws = wb.create_sheet("Details")

    ws.merge_cells("A1:D1")
    titel_cell = ws["A1"]
    titel_cell.value = f"Ueberstunden-Details  |  {fmt_datum(von_datum)} - {fmt_datum(bis_datum)}"
    titel_cell.font = Font(color=C_TITLE_FONT, bold=True, size=14)
    titel_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(color="888888", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8

    headers = ["Mitarbeiter", "Abteilung", "Datum", "Stunden", "Notiz"]
    col_count = len(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        style_header_cell(cell)
    ws.row_dimensions[4].height = 22

    row = 5
    aktuelle_abteilung = None

    for eintrag in mitarbeiter_liste:
        ma = eintrag.get("mitarbeiter", {})
        abt = ma.get("abteilung", "")
        name = ma.get("name", "")
        eintraege = eintrag.get("eintraege", [])

        if not eintraege:
            continue

        if abt != aktuelle_abteilung:
            aktuelle_abteilung = abt
            ws.merge_cells(f"A{row}:E{row}")
            abt_cell = ws.cell(row=row, column=1, value=abt)
            abt_cell.fill = PatternFill(start_color=C_ABT_BG, end_color=C_ABT_BG, fill_type="solid")
            abt_cell.font = Font(color=C_ABT_FONT, bold=True, size=10)
            abt_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            abt_cell.border = make_border()
            ws.row_dimensions[row].height = 20
            row += 1

        for e in eintraege:
            zeile = [name, abt, fmt_datum(e.get("datum")), fmt_zahl(e.get("stunden", 0)), e.get("notiz") or ""]
            for col, wert in enumerate(zeile, 1):
                cell = ws.cell(row=row, column=col, value=wert)
                style_data_cell(cell, bg=C_UEBERSTD_G if col == 4 else None, center=col in (3, 4))
            ws.row_dimensions[row].height = 16
            row += 1

        # Zwischensumme je Mitarbeiter
        ws.merge_cells(f"A{row}:C{row}")
        zs_cell = ws.cell(row=row, column=1, value=f"Summe {name}")
        style_data_cell(zs_cell, bg=C_SUMME_BG, bold=True)
        ws.cell(row=row, column=2).fill = PatternFill(start_color=C_SUMME_BG, end_color=C_SUMME_BG, fill_type="solid")
        ws.cell(row=row, column=2).border = make_border()
        ws.cell(row=row, column=3).fill = PatternFill(start_color=C_SUMME_BG, end_color=C_SUMME_BG, fill_type="solid")
        ws.cell(row=row, column=3).border = make_border()
        summe_zelle = ws.cell(row=row, column=4, value=fmt_zahl(sum(fmt_zahl(e.get("stunden", 0)) for e in eintraege)))
        style_data_cell(summe_zelle, bg=C_SUMME_BG, bold=True, center=True)
        ws.cell(row=row, column=5).fill = PatternFill(start_color=C_SUMME_BG, end_color=C_SUMME_BG, fill_type="solid")
        ws.cell(row=row, column=5).border = make_border()
        ws.row_dimensions[row].height = 18
        row += 1

    breiten = [28, 20, 14, 12, 45]
    for i, b in enumerate(breiten, 1):
        ws.column_dimensions[get_column_letter(i)].width = b

    ws.freeze_panes = "A5"


# ── Haupt ────────────────────────────────────────────────────────────────────
def create_excel(payload, output_path):
    export_data = payload.get("exportData", payload)
    mitarbeiter_liste = export_data.get("mitarbeiter", [])
    von_datum = export_data.get("vonDatum", "")
    bis_datum = export_data.get("bisDatum", "")
    gesamt_alle = export_data.get("gesamtStundenAlle", 0)

    wb = Workbook()
    schreibe_zusammenfassung(wb, mitarbeiter_liste, von_datum, bis_datum, gesamt_alle)
    schreibe_detail(wb, mitarbeiter_liste, von_datum, bis_datum)

    wb.save(output_path)
    sys.stdout.buffer.write(f"Excel erfolgreich erstellt: {output_path}\n".encode("utf-8"))


def main():
    if len(sys.argv) != 3:
        sys.stderr.write("FEHLER: Usage: python export_ueberstunden_excel.py <input.json> <output.xlsx>\n")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]

    try:
        with io.open(input_file, "r", encoding="utf-8-sig") as f:
            payload = json.load(f)
        sys.stdout.buffer.write(b"JSON gelesen\n")
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Lesen der JSON: {e}\n".encode("utf-8"))
        sys.exit(1)

    try:
        create_excel(payload, output_file)
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Erstellen der Excel: {e}\n".encode("utf-8"))
        sys.exit(1)


if __name__ == "__main__":
    main()