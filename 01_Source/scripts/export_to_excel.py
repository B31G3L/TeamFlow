#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel-Export fuer TeamFlow
Neue Struktur: Zusammenfassung + Detailtabelle
"""

import sys
import json
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("FEHLER: openpyxl nicht installiert!", file=sys.stderr)
    print("Installiere mit: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Farben ──────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F538D"
C_HEADER_FONT = "FFFFFF"
C_ABT_BG      = "2D5FA8"
C_ABT_FONT    = "FFFFFF"
C_URLAUB      = "D6F0D6"   # grün
C_KRANKHEIT   = "FAD4D4"   # rot
C_SCHULUNG    = "D4EEF7"   # blau
C_UEBERSTD    = "FFF3CD"   # gelb
C_SUMME_BG    = "E8E8E8"
C_TITLE_FONT  = "1F538D"

TYP_FARBEN = {
    "urlaub":       C_URLAUB,
    "krankheit":    C_KRANKHEIT,
    "schulung":     C_SCHULUNG,
    "ueberstunden": C_UEBERSTD,
}

TYP_LABEL = {
    "urlaub":       "Urlaub",
    "krankheit":    "Krankheit",
    "schulung":     "Schulung",
    "ueberstunden": "Ueberstunden-Abbau",
}


def make_border(thin=True):
    s = 'thin' if thin else 'medium'
    side = Side(style=s)
    return Border(left=side, right=side, top=side, bottom=side)


def fmt_datum(d):
    """YYYY-MM-DD -> DD.MM.YYYY"""
    if not d:
        return ""
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d


def fmt_zahl(v):
    if v is None:
        return 0
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except Exception:
        return 0


def style_header_cell(cell, bg=C_HEADER_BG, fg=C_HEADER_FONT, size=10, bold=True):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = make_border()


def style_data_cell(cell, bg=None, center=False, bold=False):
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(bold=bold, size=9)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center',
        wrap_text=False
    )
    cell.border = make_border()


# ── Tabellenblatt 1: Zusammenfassung ────────────────────────────────────────
def schreibe_zusammenfassung(wb, mitarbeiter_liste, von_datum, bis_datum):
    ws = wb.active
    ws.title = "Zusammenfassung"

    # Titel
    ws.merge_cells("A1:H1")
    titel_cell = ws["A1"]
    titel_cell.value = f"Abwesenheits-Uebersicht  |  {fmt_datum(von_datum)} – {fmt_datum(bis_datum)}"
    titel_cell.font = Font(color=C_TITLE_FONT, bold=True, size=14)
    titel_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Erstellt-Zeile
    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(color="888888", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 16

    # Leerzeile
    ws.row_dimensions[3].height = 8

    # Header
    headers = ["Mitarbeiter", "Abteilung", "Urlaub (T)", "Krankheit (T)", "Schulung (T)", "UE-Abbau (h)", "Eintraege", "Zeitraum"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        style_header_cell(cell)
    ws.row_dimensions[4].height = 22

    row = 5
    aktuelle_abteilung = None

    for eintrag in mitarbeiter_liste:
        ma   = eintrag.get("mitarbeiter", {})
        zus  = eintrag.get("zusammenfassung", {})
        abt  = ma.get("abteilung", "")
        name = ma.get("name", "")

        # Abteilungs-Trennzeile
        if abt != aktuelle_abteilung:
            aktuelle_abteilung = abt
            ws.merge_cells(f"A{row}:H{row}")
            abt_cell = ws.cell(row=row, column=1, value=abt)
            abt_cell.fill = PatternFill(start_color=C_ABT_BG, end_color=C_ABT_BG, fill_type="solid")
            abt_cell.font = Font(color=C_ABT_FONT, bold=True, size=10)
            abt_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            abt_cell.border = make_border()
            ws.row_dimensions[row].height = 20
            row += 1

        anzahl = len(eintrag.get("eintraege", []))

        werte = [
            name,
            abt,
            fmt_zahl(zus.get("urlaub_tage", 0)),
            fmt_zahl(zus.get("krankheit_tage", 0)),
            fmt_zahl(zus.get("schulung_tage", 0)),
            fmt_zahl(zus.get("ueberstunden_abbau", 0)),
            anzahl,
            f"{fmt_datum(von_datum)} – {fmt_datum(bis_datum)}",
        ]

        for col, wert in enumerate(werte, 1):
            cell = ws.cell(row=row, column=col, value=wert)
            center = col > 2
            style_data_cell(cell, center=center)

        ws.row_dimensions[row].height = 18
        row += 1

    # Summenzeile
    ws.merge_cells(f"A{row}:B{row}")
    summe_cell = ws.cell(row=row, column=1, value="GESAMT")
    style_data_cell(summe_cell, bg=C_SUMME_BG, bold=True, center=True)
    ws.cell(row=row, column=2).fill = PatternFill(start_color=C_SUMME_BG, end_color=C_SUMME_BG, fill_type="solid")
    ws.cell(row=row, column=2).border = make_border()

    summen = {
        3: sum(fmt_zahl(e.get("zusammenfassung", {}).get("urlaub_tage", 0))       for e in mitarbeiter_liste),
        4: sum(fmt_zahl(e.get("zusammenfassung", {}).get("krankheit_tage", 0))    for e in mitarbeiter_liste),
        5: sum(fmt_zahl(e.get("zusammenfassung", {}).get("schulung_tage", 0))     for e in mitarbeiter_liste),
        6: sum(fmt_zahl(e.get("zusammenfassung", {}).get("ueberstunden_abbau", 0)) for e in mitarbeiter_liste),
        7: sum(len(e.get("eintraege", []))                                         for e in mitarbeiter_liste),
    }
    for col, val in summen.items():
        cell = ws.cell(row=row, column=col, value=val)
        style_data_cell(cell, bg=C_SUMME_BG, bold=True, center=True)

    ws.cell(row=row, column=8).border = make_border()
    ws.row_dimensions[row].height = 20

    # Spaltenbreiten
    breiten = [28, 20, 12, 14, 13, 14, 10, 22]
    for i, b in enumerate(breiten, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = b

    ws.freeze_panes = "A5"


# ── Tabellenblatt 2: Detailtabelle ──────────────────────────────────────────
def schreibe_detail(wb, mitarbeiter_liste, von_datum, bis_datum):
    ws = wb.create_sheet("Details")

    # Titel
    ws.merge_cells("A1:G1")
    titel_cell = ws["A1"]
    titel_cell.value = f"Abwesenheits-Details  |  {fmt_datum(von_datum)} – {fmt_datum(bis_datum)}"
    titel_cell.font = Font(color=C_TITLE_FONT, bold=True, size=14)
    titel_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(color="888888", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8

    headers = ["Mitarbeiter", "Abteilung", "Typ", "Von", "Bis", "Wert", "Notiz / Titel"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        style_header_cell(cell)
    ws.row_dimensions[4].height = 22

    row = 5
    aktuelle_abteilung = None

    for eintrag in mitarbeiter_liste:
        ma       = eintrag.get("mitarbeiter", {})
        eintraege = eintrag.get("eintraege", [])
        abt      = ma.get("abteilung", "")
        name     = ma.get("name", "")

        if not eintraege:
            continue

        # Abteilungs-Trennzeile
        if abt != aktuelle_abteilung:
            aktuelle_abteilung = abt
            ws.merge_cells(f"A{row}:G{row}")
            abt_cell = ws.cell(row=row, column=1, value=abt)
            abt_cell.fill = PatternFill(start_color=C_ABT_BG, end_color=C_ABT_BG, fill_type="solid")
            abt_cell.font = Font(color=C_ABT_FONT, bold=True, size=10)
            abt_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            abt_cell.border = make_border()
            ws.row_dimensions[row].height = 20
            row += 1

        for e in eintraege:
            typ   = e.get("typ", "")
            farbe = TYP_FARBEN.get(typ, "FFFFFF")
            label = TYP_LABEL.get(typ, typ)

            wert = fmt_zahl(e.get("wert", 0))
            einheit = "h" if typ == "ueberstunden" else "T"
            wert_str = f"{wert} {einheit}"

            notiz = e.get("notiz") or e.get("titel") or ""

            zeile = [name, abt, label, fmt_datum(e.get("von_datum")), fmt_datum(e.get("bis_datum")), wert_str, notiz]

            for col, val in enumerate(zeile, 1):
                cell = ws.cell(row=row, column=col, value=val)
                center = col in (3, 4, 5, 6)
                style_data_cell(cell, bg=farbe if col > 2 else None, center=center)

            ws.row_dimensions[row].height = 16
            row += 1

    # Spaltenbreiten
    breiten = [28, 20, 18, 12, 12, 10, 35]
    for i, b in enumerate(breiten, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = b

    ws.freeze_panes = "A5"


# ── Legende ─────────────────────────────────────────────────────────────────
def schreibe_legende(wb):
    ws = wb.create_sheet("Legende")

    ws["A1"].value = "Farbcode"
    style_header_cell(ws["A1"])
    ws["B1"].value = "Bedeutung"
    style_header_cell(ws["B1"])
    ws.row_dimensions[1].height = 20

    legende = [
        (C_URLAUB,    "Urlaub (Tage)"),
        (C_KRANKHEIT, "Krankheit (Tage)"),
        (C_SCHULUNG,  "Schulung (Tage)"),
        (C_UEBERSTD,  "Ueberstunden-Abbau (Stunden)"),
        (C_ABT_BG,    "Abteilung"),
        (C_SUMME_BG,  "Summenzeile"),
    ]

    for i, (farbe, text) in enumerate(legende, 2):
        cell_a = ws.cell(row=i, column=1, value="")
        cell_a.fill = PatternFill(start_color=farbe, end_color=farbe, fill_type="solid")
        cell_a.border = make_border()
        cell_b = ws.cell(row=i, column=2, value=text)
        style_data_cell(cell_b)
        ws.row_dimensions[i].height = 18

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30


# ── Haupt ────────────────────────────────────────────────────────────────────
def create_excel(payload, output_path):
    export_data = payload.get("exportData", payload)
    mitarbeiter_liste = export_data.get("mitarbeiter", [])
    von_datum = export_data.get("vonDatum", "")
    bis_datum = export_data.get("bisDatum", "")

    wb = Workbook()

    schreibe_zusammenfassung(wb, mitarbeiter_liste, von_datum, bis_datum)
    schreibe_detail(wb, mitarbeiter_liste, von_datum, bis_datum)
    schreibe_legende(wb)

    wb.save(output_path)
    sys.stdout.buffer.write(f"Excel erfolgreich erstellt: {output_path}\n".encode("utf-8"))


def main():
    if len(sys.argv) != 3:
        sys.stderr.write("FEHLER: Usage: python export_to_excel.py <input.json> <output.xlsx>\n")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]

    try:
        with io.open(input_file, "r", encoding="utf-8-sig") as f:
            payload = json.load(f)
        mitarbeiter_liste = payload.get("exportData", payload).get("mitarbeiter", payload)
        anzahl = len(mitarbeiter_liste) if isinstance(mitarbeiter_liste, list) else "?"
        sys.stdout.buffer.write(f"JSON gelesen: {anzahl} Mitarbeiter\n".encode("utf-8"))
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Lesen der JSON: {e}\n".encode("utf-8"))
        sys.exit(1)

    try:
        create_excel(payload, output_file)
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Erstellen der Excel: {e}\n".encode("utf-8"))
        sys.exit(1)


if __name__ == "__main__":
    main()